import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment

from product import Product
from src.goods_ids import rosel_products


class PagesParser:
    def __init__(self, directory):
        self.platform = None
        self.dir = directory
        self.date = directory.split('/')[1]
        self.soup = None
        self.cp = None
        self.html_file = None
        self.workbook = Workbook()

    def run(self):
        files = sorted(os.listdir(self.dir))
        self.initiate_workbook()
        platform = self.platform
        for n, f in enumerate(files):
            filename = f"{self.dir}/{f}"
            self.html_file = f
            # print(f'{platform} - №{n + 1:03} - Open {filename}')
            with open(filename, 'r', encoding='utf8') as read_file:
                self.set_product(int(f.split('_')[-1].split('.')[0]))
                self.soup = BeautifulSoup(read_file, 'lxml')
                self.parse()
                self.data_out()
                # if n > 30:
                #     break
        self.workbook.save(f'results/{platform}-{self.date}.xlsx')

    def set_product(self, art):
        prod = rosel_products[art]
        shop_ids = {'wb': prod['id WB'], 'oz': prod['id OZ']}
        shop_urls = {'wb': prod['url WB'], 'oz': prod['url OZ']}
        self.cp = Product()
        self.cp.rosel_id = art
        self.cp.shop_id = shop_ids[self.platform]
        self.cp.name = prod['Name']
        self.cp.url = shop_urls[self.platform]

    def parse(self):
        pass

    def data_out(self):
        write_data = self.cp.xlsx_line()
        self.workbook.active.append(write_data)

    def initiate_workbook(self):
        titles = ['ID', 'Shop ID', 'Наименование', 'Рейтинг', 'Цель из Ф4', 'Необходимо отзывов с 5*', 'Количество',
                  '5*', '4*', '3*', '2*', '1*', 'Статус', 'Price', 'url']
        self.workbook.create_sheet(self.platform)
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        self.workbook.active.append(titles)
        self.set_workbook_dimensions()

    def set_workbook_dimensions(self):
        sw = self.workbook.active
        sw.column_dimensions['A'].width = 6
        sw.column_dimensions['B'].width = 18
        sw.column_dimensions['C'].width = 30
        mark_value = 6
        sw.column_dimensions['E'].width = mark_value
        sw.column_dimensions['E'].width = mark_value
        sw.column_dimensions['H'].width = mark_value
        sw.column_dimensions['I'].width = mark_value
        sw.column_dimensions['J'].width = mark_value
        sw.column_dimensions['K'].width = mark_value
        sw.column_dimensions['L'].width = mark_value


class ParserOz(PagesParser):
    def __init__(self, directory):
        super().__init__(directory)
        self.platform = 'oz'

    def parse(self):
        cp = self.cp
        soup = self.soup
        try:
            reviews_block = soup.find('div', attrs={"data-widget": "webReviewProductScore"}).a['title']
            cp.quantity = int(reviews_block.split()[0])
        except TypeError:
            pass
        if cp.quantity:
            try:
                votes_stars_parent = soup.find(text='5 звезд').parent.parent
                cp.rating = float(votes_stars_parent.parent.parent.find('span').text.split()[0])
                votes_class = votes_stars_parent.find_all('div')[4]['class'][0]
                votes = [int(vote.text) for vote in soup.find_all('div', class_=votes_class)]
                cp.votes = dict(zip(['5*', '4*', '3*', '2*', '1*'], votes))
            except Exception as ex:
                print(f'{self.html_file} - {cp.shop_id}: rating error - {ex}')
        delivery = soup.find('h2', text='Информация о доставке')
        if delivery:
            cp.status = delivery.parent.find(text='В наличии')
        self.get_price()

    def get_price(self):
        try:
            prb = self.soup.find('div', attrs={"data-widget": "webPrice"}).div.find_all('div', recursive=False)[1]
            self.cp.price = prb.div.span.text.strip()
            return
        except Exception as ex:
            try:
                prb = self.soup.find('div', attrs={"data-widget": "webPrice"}).div.div.div
                self.cp.price = prb.span.text.strip()
            except Exception as ex:
                print(f'{self.html_file} - {self.cp.shop_id}: price error - {ex}')


class ParserWb(PagesParser):
    def __init__(self, directory):
        super().__init__(directory)
        self.platform = 'wb'

    def parse(self):
        cp = self.cp
        soup = self.soup
        try:
            cp.quantity = soup.find('span', class_='same-part-kt__count-review').text.strip().split()[0]
        except AttributeError:
            cp.status = 'ERROR'
            return
        self.get_status()
        if cp.quantity == '0':
            return
        self.get_rating()

    def get_status(self):
        cp = self.cp
        try:
            status = self.soup.find('span', class_='same-part-kt__delivery-info').text.strip()
            cp.status = status
        except AttributeError:
            try:
                status = self.soup.find('span', class_='sold-out-product__text').text.split()
                cp.status = ' '.join(status)
            except AttributeError:
                cp.status = None

    def get_rating(self):
        cp = self.cp
        try:
            html_rating = self.soup.find('div', class_='user-scores__rating')
            votes = [vote.text for vote in html_rating.find_all('span', class_='progress-lines__percent')]
            cp.votes = dict(zip(['5*', '4*', '3*', '2*', '1*'], votes))
            cp.rating = float(html_rating.find('span', class_='user-scores__score').text)
            for key, value in cp.votes.items():
                votes = int(round((int(value[:-1]) * int(cp.quantity)) / 100, 0))
                cp.votes[key] = votes
        except Exception as ex:
            cp.rating = -1
            print(f'{self.html_file} - {cp.rosel_id}: rating error - {ex}')
