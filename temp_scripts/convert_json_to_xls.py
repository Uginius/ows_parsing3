import json
import math

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

from temp_scripts.wb_prices import prices_WB

with open('json_data/2022-07-05_wb.json', 'r', encoding='utf8') as json_file:
    json_dict = json.load(json_file)

workbook = load_workbook('results/wb-2022-07-04.xlsx')
ws = workbook.active
for n, row in enumerate(ws, start=1):
    if n == 1:
        continue
    rosel_id = str(row[0].value)
    merch = json_dict.get(rosel_id)
    if merch:
        rating = merch['Рейтинг']
        qt = merch['Количество отзывов']
    else:
        rating = 0
        qt = 0
    row[6].value = qt
    if rating > 4 and qt > 1:
        need = 0
    else:
        need = math.ceil(qt * abs(4 - rating)) if rating else 2
    rating_cell = row[3]
    rating_cell.value = rating
    if rating_cell.value == 0:
        rating_cell.fill = PatternFill("solid", fgColor='ffe4e1')
    elif rating_cell.value < 4:
        rating_cell.fill = PatternFill("solid", fgColor='E6B8B7')
    elif rating_cell.value > 4:
        rating_cell.fill = PatternFill("solid", fgColor='7CFC00')
    need_votes_cell = row[5]
    need_votes_cell.value = need
    if need_votes_cell.value > 0:
        need_votes_cell.fill = PatternFill("solid", fgColor='ffcd75')
    try:
        price = prices_WB.get(int(rosel_id))
    except ValueError:
        price = None
    if price:
        row[7].value = price

workbook.save(f"results/rating.xlsx")
